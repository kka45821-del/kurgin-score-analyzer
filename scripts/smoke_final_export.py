from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_output.final_export import build_final_excel_bytes, export_final_excel
from scripts.smoke_final_excel_contract import validate_workbook


def _sample_rows() -> list[tuple[dict, dict]]:
    source = {
        "row_id": 1,
        "stone_id": "KRG-EXPORT-001",
        "Report #": "LG-EXPORT-001",
        "Lab": "IGI",
        "Stock #": "EXPORT-STOCK-001",
        "Shape": "Round",
        "Weight": 1.33,
        "Color": "E",
        "Clarity": "VS1",
        "Measurements": "6.95x6.98x4.29",
        "min_diameter": 6.95,
        "max_diameter": 6.98,
        "avg_diameter": 6.965,
        "depth_mm": 4.29,
        "TablePercent": 57.0,
        "DepthPercent": 61.7,
        "CrownAngle": 34.6,
        "CrownPercent": 15.0,
        "PavilionAngle": 40.8,
        "PavilionPercent": 43.0,
        "GirdlePercent": 3.5,
        "Fluorescence": "None",
        "Cut": "Excellent",
        "Polish": "Excellent",
        "Symmetry": "Excellent",
    }
    analysis = {
        "final_score": 89.8,
        "score_band": "high",
        "score_label": "High",
        "summary": "Export smoke public-safe summary.",
        "tags": ["export", "smoke"],
        "data_completeness": "complete",
        "validation_status": "ready",
        "admin_review_required": False,
        "class": "controlled",
        "recommendation": "Ready for export smoke.",
        "warnings": [],
        "limitations": ["Not a certificate", "Not an appraisal"],
        "report_quality": "good",
        "visual_check": False,
        "critical_risk": False,
        "triple_score": 91.0,
        "structure_modifier": 0.99,
        "diagnostics": "internal diagnostics export smoke",
        "breakdown": "internal breakdown export smoke",
        "engine_version": "export-smoke",
        "calculation_status": "mock",
        "internal_warnings": "internal only",
    }
    return [(source, analysis)]


def _assert_exported_workbook(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        assert "PUBLIC_CATALOG_IMPORT" in workbook.sheetnames, workbook.sheetnames
    finally:
        workbook.close()


def main() -> int:
    rows = _sample_rows()
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = Path(tmpdir) / "nested" / "final_export_smoke.xlsx"
        exported_path = export_final_excel(rows, output_path)
        assert exported_path == output_path, exported_path
        assert exported_path.exists(), exported_path

        errors = validate_workbook(exported_path)
        if errors:
            print("Final export smoke failed contract validation:")
            for error in errors:
                print(f"- {error}")
            return 1
        _assert_exported_workbook(exported_path)

    excel_bytes = build_final_excel_bytes(rows)
    assert isinstance(excel_bytes, bytes), type(excel_bytes)
    assert len(excel_bytes) > 0, "Expected non-empty Excel bytes"

    print("Final export smoke passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
