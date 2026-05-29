from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

REQUIRED_SHEETS = [
    "PUBLIC_CATALOG_IMPORT",
    "ANALYZER_RESULTS",
    "VALIDATION_ERRORS",
    "INTERNAL_DIAGNOSTICS",
    "README_SCHEMA",
]

PUBLIC_CATALOG_IMPORT_COLUMNS = [
    "stone_id",
    "report_number",
    "laboratory",
    "stock_number",
    "shape",
    "carat",
    "color",
    "clarity",
    "measurements",
    "min_diameter",
    "max_diameter",
    "avg_diameter",
    "depth_mm",
    "table_pct",
    "depth_pct",
    "crown_angle",
    "crown_height_pct",
    "pavilion_angle",
    "pavilion_depth_pct",
    "girdle_pct",
    "fluorescence",
    "cut",
    "polish",
    "symmetry",
    "kurgin_score",
    "score_band",
    "public_score_label",
    "public_summary",
    "public_tags",
    "data_quality_status",
    "validation_status",
    "admin_review_required",
    "show_in_catalog",
    "price_status",
    "public_action",
    "availability_status",
]

FORBIDDEN_PUBLIC_COLUMNS = [
    "raw_formula",
    "weights",
    "formula_thresholds",
    "penalty_breakdown",
    "raw_diagnostics",
    "debug_trace",
    "internal_formula_config",
    "coefficient_formula",
    "raw_engine_output",
    "breakdown",
    "certificate_claim",
    "appraisal_claim",
    "guaranteed_price_effect",
    "order_effect",
    "reserve_effect",
    "payment_effect",
]


def _header_values(sheet) -> list[str]:
    values = []
    for cell in sheet[1]:
        if cell.value is None:
            continue
        values.append(str(cell.value).strip())
    return values


def validate_workbook(path: str | Path) -> list[str]:
    """Validate a final Analyzer Excel workbook against the v0.1 contract."""
    errors: list[str] = []
    workbook_path = Path(path)

    if not workbook_path.exists():
        return [f"Workbook not found: {workbook_path}"]

    try:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:
        return [f"Workbook cannot be opened: {exc}"]

    sheet_names = set(workbook.sheetnames)
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in sheet_names:
            errors.append(f"Missing required sheet: {sheet_name}")

    if "PUBLIC_CATALOG_IMPORT" in sheet_names:
        public_sheet = workbook["PUBLIC_CATALOG_IMPORT"]
        headers = _header_values(public_sheet)
        header_set = set(headers)

        for column in PUBLIC_CATALOG_IMPORT_COLUMNS:
            if column not in header_set:
                errors.append(f"PUBLIC_CATALOG_IMPORT missing required column: {column}")

        for column in FORBIDDEN_PUBLIC_COLUMNS:
            if column in header_set:
                errors.append(f"PUBLIC_CATALOG_IMPORT contains forbidden public column: {column}")

    if "README_SCHEMA" not in sheet_names:
        errors.append("README_SCHEMA sheet is required")

    try:
        workbook.close()
    except Exception:
        pass

    return errors


def _create_minimal_valid_workbook(path: Path) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = "PUBLIC_CATALOG_IMPORT"
    default_sheet.append(PUBLIC_CATALOG_IMPORT_COLUMNS)

    for sheet_name in REQUIRED_SHEETS:
        if sheet_name == "PUBLIC_CATALOG_IMPORT":
            continue
        sheet = workbook.create_sheet(sheet_name)
        if sheet_name == "README_SCHEMA":
            sheet.append(["contract_name", "contract_version", "admin_import_sheet"])
            sheet.append(["FINAL_ANALYZER_EXCEL_OUTPUT_CONTRACT", "0.1", "PUBLIC_CATALOG_IMPORT"])
        else:
            sheet.append(["row_id", "stone_id", "report_number"])

    workbook.save(path)


def _self_test() -> int:
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_path = Path(tmpdir) / "minimal_final_analyzer_contract.xlsx"
        _create_minimal_valid_workbook(workbook_path)
        errors = validate_workbook(workbook_path)
        if errors:
            print("Self-test failed:")
            for error in errors:
                print(f"- {error}")
            return 1
    print("Final Analyzer Excel contract self-test passed.")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        return _self_test()

    errors = validate_workbook(argv[1])
    if errors:
        print("Final Analyzer Excel contract validation failed:")
        for error in errors:
            print(f"- {error}")
        return 1

    print("Final Analyzer Excel contract validation passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
