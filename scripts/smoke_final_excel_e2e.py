from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_output.final_mapper_bridge import build_final_workbook_from_analysis_rows
from scripts.smoke_final_excel_contract import FORBIDDEN_PUBLIC_COLUMNS, validate_workbook

EXPECTED_SOURCE_ROW_COUNT = 4


def _base_source(row_id: int, stone_id: str, report_number: str) -> dict:
    return {
        "row_id": row_id,
        "stone_id": stone_id,
        "Report #": report_number,
        "Lab": "IGI",
        "Stock #": f"E2E-STOCK-{row_id:03d}",
        "Shape": "Round",
        "Weight": 1.25,
        "Color": "E",
        "Clarity": "VS1",
        "Measurements": "6.85x6.88x4.23",
        "min_diameter": 6.85,
        "max_diameter": 6.88,
        "avg_diameter": 6.865,
        "depth_mm": 4.23,
        "TablePercent": 57.0,
        "DepthPercent": 61.8,
        "CrownAngle": 34.7,
        "CrownPercent": 15.0,
        "PavilionAngle": 40.8,
        "PavilionPercent": 43.0,
        "GirdlePercent": 3.5,
        "Fluorescence": "None",
        "Cut": "Excellent",
        "Polish": "Excellent",
        "Symmetry": "Excellent",
    }


def _analysis(
    *,
    score: float | None,
    validation_status: str,
    admin_review_required: bool,
    diagnostics_suffix: str,
) -> dict:
    analysis = {
        "score_band": "high" if score is not None else "",
        "score_label": "High" if score is not None else "",
        "summary": f"E2E public-safe summary {diagnostics_suffix}.",
        "tags": ["e2e", diagnostics_suffix],
        "data_completeness": "complete" if validation_status == "ready" else "incomplete",
        "validation_status": validation_status,
        "admin_review_required": admin_review_required,
        "class": "controlled",
        "recommendation": "Review according to validation status.",
        "warnings": [] if validation_status == "ready" else ["Review required"],
        "limitations": ["Not a certificate", "Not an appraisal"],
        "report_quality": "good" if validation_status == "ready" else "limited",
        "visual_check": False,
        "critical_risk": False,
        "triple_score": 91.0,
        "structure_modifier": 0.99,
        "diagnostics": f"internal diagnostics {diagnostics_suffix}",
        "breakdown": f"internal breakdown {diagnostics_suffix}",
        "engine_version": "e2e-smoke",
        "calculation_status": "mock",
        "internal_warnings": f"internal warning {diagnostics_suffix}",
        "raw_formula": "must not reach public catalog",
        "weights": {"secret": 1},
        "debug_trace": ["secret"],
        "formula_thresholds": {"secret": 1},
    }
    if score is not None:
        analysis["final_score"] = score
    return analysis


def _sample_pairs() -> list[tuple[dict, dict]]:
    ready_source = _base_source(1, "KRG-E2E-001", "LG-E2E-001")
    ready_analysis = _analysis(
        score=90.1,
        validation_status="ready",
        admin_review_required=False,
        diagnostics_suffix="ready",
    )

    incomplete_source = _base_source(2, "KRG-E2E-002", "LG-E2E-002")
    incomplete_source.pop("TablePercent")
    incomplete_source.pop("DepthPercent")
    incomplete_analysis = _analysis(
        score=76.0,
        validation_status="review_required",
        admin_review_required=True,
        diagnostics_suffix="incomplete",
    )

    missing_score_source = _base_source(3, "KRG-E2E-003", "LG-E2E-003")
    missing_score_analysis = _analysis(
        score=None,
        validation_status="ready",
        admin_review_required=False,
        diagnostics_suffix="missing-score",
    )

    missing_identity_source = _base_source(4, "", "")
    missing_identity_analysis = _analysis(
        score=81.0,
        validation_status="blocked",
        admin_review_required=True,
        diagnostics_suffix="missing-identity",
    )

    return [
        (ready_source, ready_analysis),
        (incomplete_source, incomplete_analysis),
        (missing_score_source, missing_score_analysis),
        (missing_identity_source, missing_identity_analysis),
    ]


def _headers(sheet) -> list[str]:
    return [str(cell.value).strip() for cell in sheet[1] if cell.value is not None]


def _records(sheet) -> list[dict]:
    headers = _headers(sheet)
    records: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        records.append({headers[index]: value for index, value in enumerate(row[: len(headers)])})
    return records


def _assert_workbook_contents(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        public_sheet = workbook["PUBLIC_CATALOG_IMPORT"]
        validation_sheet = workbook["VALIDATION_ERRORS"]
        diagnostics_sheet = workbook["INTERNAL_DIAGNOSTICS"]

        public_headers = set(_headers(public_sheet))
        forbidden_present = sorted(public_headers.intersection(FORBIDDEN_PUBLIC_COLUMNS))
        assert not forbidden_present, f"Forbidden public columns found: {forbidden_present}"

        public_records = _records(public_sheet)
        validation_records = _records(validation_sheet)
        diagnostics_records = _records(diagnostics_sheet)

        assert len(public_records) == EXPECTED_SOURCE_ROW_COUNT, public_records
        assert len(diagnostics_records) == EXPECTED_SOURCE_ROW_COUNT, diagnostics_records
        assert any(record.get("error_code") == "validation_status_not_ready" for record in validation_records), validation_records
        assert any(record.get("error_code") == "score_missing" for record in validation_records), validation_records
        assert any(record.get("error_code") == "missing_required_identity" for record in validation_records), validation_records
        assert any(record.get("blocks_catalog_import") == "yes" for record in validation_records), validation_records
        assert all("internal diagnostics" in str(record.get("diagnostics", "")) for record in diagnostics_records), diagnostics_records
        assert all("internal breakdown" in str(record.get("breakdown", "")) for record in diagnostics_records), diagnostics_records
    finally:
        workbook.close()


def main() -> int:
    workbook = build_final_workbook_from_analysis_rows(_sample_pairs())
    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_path = Path(tmpdir) / "final_excel_e2e_smoke.xlsx"
        workbook.save(workbook_path)
        errors = validate_workbook(workbook_path)
        if errors:
            print("Final Excel E2E smoke failed contract validation:")
            for error in errors:
                print(f"- {error}")
            return 1
        _assert_workbook_contents(workbook_path)

    print("Final Excel E2E smoke passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
