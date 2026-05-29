from __future__ import annotations

import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from excel_output.final_workbook_builder import build_final_workbook
from scripts.smoke_final_excel_contract import PUBLIC_CATALOG_IMPORT_COLUMNS, validate_workbook


def _public_row() -> dict:
    return {
        "stone_id": "KRG-TEST-001",
        "report_number": "LG123456789",
        "laboratory": "IGI",
        "stock_number": "STOCK-001",
        "shape": "Round",
        "carat": 1.25,
        "color": "E",
        "clarity": "VS1",
        "measurements": "6.85x6.88x4.23",
        "avg_diameter": 6.865,
        "table_pct": 57.0,
        "depth_pct": 61.8,
        "crown_angle": 34.7,
        "pavilion_angle": 40.8,
        "crown_height_pct": 15.0,
        "pavilion_depth_pct": 43.0,
        "girdle_pct": 3.5,
        "fluorescence": "None",
        "cut": "Excellent",
        "polish": "Excellent",
        "symmetry": "Excellent",
        "kurgin_score": 88.5,
        "score_band": "high",
        "public_score_label": "High",
        "public_summary": "Controlled public-safe summary.",
        "public_tags": "balanced,reviewed",
        "data_quality_status": "complete",
        "validation_status": "ready",
        "admin_review_required": False,
        "show_in_catalog": True,
        "price_status": "request_price",
        "public_action": "request_conditions",
        "availability_status": "available",
    }


def _analyzer_row() -> dict:
    return {
        "row_id": 1,
        "stone_id": "KRG-TEST-001",
        "report_number": "LG123456789",
        "kurgin_score": 88.5,
        "score_band": "high",
        "class": "controlled",
        "tags": "balanced,reviewed",
        "short_conclusion": "Controlled Analyzer result row.",
        "recommendation": "Admin review not required.",
        "warnings": "",
        "limitations": "Not a certificate; not an appraisal.",
        "data_completeness": "complete",
        "report_quality": "good",
        "visual_check": False,
        "critical_risk": False,
    }


def _validation_row() -> dict:
    return {
        "row_id": 1,
        "stone_id": "KRG-TEST-001",
        "report_number": "LG123456789",
        "field": "public_summary",
        "error_code": "review_note",
        "severity": "warning",
        "message": "Example warning row.",
        "recommended_action": "Confirm public wording before publication.",
        "blocks_catalog_import": "no",
    }


def _diagnostics_row() -> dict:
    return {
        "row_id": 1,
        "stone_id": "KRG-TEST-001",
        "report_number": "LG123456789",
        "triple_score": 91.0,
        "structure_modifier": 0.98,
        "diagnostics": "internal-only diagnostics placeholder",
        "breakdown": "internal-only breakdown placeholder",
        "engine_version": "pending",
        "calculation_status": "mock",
        "internal_warnings": "internal-only",
    }


def _assert_public_sheet_shape(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["PUBLIC_CATALOG_IMPORT"]
        headers = [cell.value for cell in sheet[1]]
        assert headers == PUBLIC_CATALOG_IMPORT_COLUMNS, "PUBLIC_CATALOG_IMPORT column order changed"
        assert sheet.max_row == 2, f"Expected exactly 1 data row, got {sheet.max_row - 1}"
    finally:
        workbook.close()


def _assert_forbidden_public_column_rejected() -> None:
    bad_row = _public_row()
    bad_row["raw_formula"] = "should not be accepted"
    try:
        build_final_workbook(public_rows=[bad_row])
    except ValueError as exc:
        assert "raw_formula" in str(exc), exc
        return
    raise AssertionError("Forbidden public column did not raise ValueError")


def main() -> int:
    workbook = build_final_workbook(
        public_rows=[_public_row()],
        analyzer_rows=[_analyzer_row()],
        validation_rows=[_validation_row()],
        diagnostics_rows=[_diagnostics_row()],
    )

    with tempfile.TemporaryDirectory() as tmpdir:
        workbook_path = Path(tmpdir) / "final_workbook_rows_smoke.xlsx"
        workbook.save(workbook_path)
        errors = validate_workbook(workbook_path)
        if errors:
            print("Final workbook rows smoke failed contract validation:")
            for error in errors:
                print(f"- {error}")
            return 1
        _assert_public_sheet_shape(workbook_path)

    _assert_forbidden_public_column_rejected()
    print("Final workbook rows smoke passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
