from __future__ import annotations

from io import BytesIO
from zipfile import ZipFile

import pandas as pd
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app
from kurgin_core import analyze_dataframe, analyze_stone, export_analysis_package, export_excel


EXPECTED_SHEETS = ["Results", "Details", "Issues", "System"]

RESULTS_REQUIRED_COLUMNS = [
    "Stock #",
    "Availability",
    "Report #",
    "Lab",
    "Shape",
    "Weight",
    "Color",
    "Clarity",
    "Measurements",
    "Kurgin Score",
    "Verdict Local",
    "score_band_label_ru",
    "tags_all",
    "tag1",
    "tag2",
    "tag3",
    "tag4",
    "tag5",
    "tag6",
    "interpretation_short_ru",
    "recommendation_ru",
    "warning_ru",
    "Data Completeness %",
    "Report Quality Status",
    "platform_import_status",
    "recommended_pdf_priority",
    "PDF Report Status",
    "PDF Report File",
    "PDF Generation Mode",
    "Upload Quality Status",
    "Geometry Status",
    "Column Recognition Status",
    "Calculation Status",
]

DETAILS_REQUIRED_COLUMNS = [
    "KURGIN Import ID",
    "Stone Title",
    "Identification Line",
    "Report #",
    "Lab",
    "Shape",
    "Kurgin Score",
    "Verdict",
    "Verdict Local",
    "Triple Score",
    "Structure Modifier",
    "Visual Check",
    "Critical Risk",
    "KURGIN Report ID",
    "PDF Report Status",
    "PDF Report File",
    "Report Template Version",
    "Formula Output Version",
    "Engine Version",
    "Calculation Status",
    "Breakdown",
]

ISSUES_REQUIRED_COLUMNS = [
    "Issue Type",
    "Stock #",
    "Report #",
    "Lab",
    "Shape",
    "Calculation Status",
    "Problem",
    "Validation Errors",
    "Missing Fields",
    "recommendation_ru",
    "PDF Report Status",
]

SYSTEM_REQUIRED_FIELDS = [
    "Engine Version",
    "Report Template Version",
    "Formula Output Version",
    "Formula Mode",
    "Supported Shapes",
]

ANALYZE_RESULT_REQUIRED_FIELDS = [
    "Calculation Status",
    "Kurgin Score",
    "Verdict",
    "Verdict Local",
    "score_band_label_ru",
    "tags_all",
    "tag1",
    "tag2",
    "tag3",
    "tag4",
    "tag5",
    "tag6",
    "interpretation_short_ru",
    "recommendation_ru",
    "warning_ru",
]

PLATFORM_CARD_REQUIRED_KEYS = [
    "kind",
    "schema_version",
    "status",
    "identity",
    "stone",
    "score",
    "tags",
    "summary",
    "measurements",
    "quality",
    "report",
    "version",
    "actions",
]

PLATFORM_FORBIDDEN_KEYS = {
    "breakdown",
    "diagnostics",
    "triple_score",
    "structure_modifier",
    "Triple Score",
    "Structure Modifier",
    "Breakdown",
    "Nailhead",
    "Fisheye",
    "Fire Loss",
    "Depth Dev",
    "Crown Dev",
    "Pavilion Dev",
    "Balance Err",
    "Girdle Penalty",
}


def sample_ok_stone():
    return {
        "Stock #": "SCHEMA-OK",
        "Availability": "Available",
        "Report #": "SCHEMA-OK-001",
        "Lab": "IGI",
        "Shape": "ROUND",
        "Weight": 1.0,
        "Color": "F",
        "Clarity": "VS1",
        "Cut": "Excellent",
        "Polish": "Excellent",
        "Symmetry": "Excellent",
        "Fluorescence Intensity": "None",
        "Measurements": "6.430x6.470x3.970",
        "CrownAngle": 34.5,
        "PavilionAngle": 40.8,
        "TablePercent": 56,
        "DepthPercent": 61.5,
        "CrownPercent": 15,
        "PavilionPercent": 43,
        "GirdlePercent": 3.5,
    }


def sample_unsupported_stone():
    row = sample_ok_stone()
    row.update({
        "Stock #": "SCHEMA-UNSUPPORTED",
        "Report #": "SCHEMA-UNSUPPORTED-001",
        "Shape": "OVAL",
    })
    return row


def assert_contains(container, required, label):
    missing = [item for item in required if item not in container]
    assert not missing, f"{label} missing required items: {missing}"


def sheet_headers(workbook, sheet_name):
    sheet = workbook[sheet_name]
    return [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value]


def has_key_recursive(value, forbidden):
    if isinstance(value, dict):
        for key, child in value.items():
            if key in forbidden:
                return True, key
            found, found_key = has_key_recursive(child, forbidden)
            if found:
                return True, found_key
    elif isinstance(value, list):
        for child in value:
            found, found_key = has_key_recursive(child, forbidden)
            if found:
                return True, found_key
    return False, None


def test_excel_workbook_schema():
    df = pd.DataFrame([sample_ok_stone(), sample_unsupported_stone()])
    batch = analyze_dataframe(df, language="RU")
    workbook_bytes = export_excel(batch)
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    assert workbook.sheetnames == EXPECTED_SHEETS, workbook.sheetnames

    results_headers = sheet_headers(workbook, "Results")
    details_headers = sheet_headers(workbook, "Details")
    issues_headers = sheet_headers(workbook, "Issues")
    system_headers = sheet_headers(workbook, "System")

    assert_contains(results_headers, RESULTS_REQUIRED_COLUMNS, "Results")
    assert_contains(details_headers, DETAILS_REQUIRED_COLUMNS, "Details")
    assert_contains(issues_headers, ISSUES_REQUIRED_COLUMNS, "Issues")
    assert system_headers == ["Section", "Field", "Value", "Detail"], system_headers

    system_sheet = workbook["System"]
    system_fields = [row[1] for row in system_sheet.iter_rows(min_row=2, values_only=True) if row and len(row) > 1]
    assert_contains(system_fields, SYSTEM_REQUIRED_FIELDS, "System fields")


def test_analysis_package_schema():
    df = pd.DataFrame([sample_ok_stone(), sample_unsupported_stone()])
    batch = analyze_dataframe(df, language="RU")
    package = export_analysis_package(batch, pdf_mode="all_ok")

    with ZipFile(BytesIO(package)) as zf:
        names = sorted(zf.namelist())
        assert "kurgin_score_result.xlsx" in names, names
        pdf_names = [name for name in names if name.startswith("reports/") and name.endswith("_KURGIN_Report.pdf")]
        assert pdf_names == ["reports/SCHEMA-OK-001_KURGIN_Report.pdf"], names


def test_api_schema_shapes():
    client = TestClient(app)

    health = client.get("/v1/health")
    assert health.status_code == 200, health.text
    health_json = health.json()
    assert health_json["status"] == "ok"
    assert_contains(health_json["version"].keys(), [
        "core_sdk_version",
        "engine_version",
        "formula_mode",
        "default_report_level",
        "api_version",
        "service_name",
        "environment",
    ], "health.version")

    ready = client.get("/v1/ready")
    assert ready.status_code == 200, ready.text
    ready_json = ready.json()
    assert ready_json["status"] in {"ready", "not_ready"}
    assert "calculation_status" in ready_json
    assert "version" in ready_json

    payload = {"language": "RU", "stone": sample_ok_stone()}
    analyzed = client.post("/v1/analyze/stone", json=payload)
    assert analyzed.status_code == 200, analyzed.text
    analyzed_json = analyzed.json()
    assert analyzed_json["status"] == "OK"
    assert "result" in analyzed_json
    assert "version" in analyzed_json
    assert_contains(analyzed_json["result"].keys(), ANALYZE_RESULT_REQUIRED_FIELDS, "analyze.result")

    batch = client.post("/v1/analyze/batch/json", json={
        "language": "RU",
        "rows": [sample_ok_stone(), sample_unsupported_stone()],
        "include_records": True,
    })
    assert batch.status_code == 200, batch.text
    batch_json = batch.json()
    assert batch_json["status"] == "OK"
    assert batch_json["rows_total"] == 2
    assert batch_json["rows_ok"] == 1
    assert batch_json["rows_issues"] == 1
    assert isinstance(batch_json["records"], list)
    assert "version" in batch_json

    card = client.post("/v1/platform/stone-card?include_raw=false&include_experimental=false", json=payload)
    assert card.status_code == 200, card.text
    card_json = card.json()
    assert card_json["status"] == "OK"
    assert "card" in card_json
    assert "result" not in card_json
    assert_contains(card_json["card"].keys(), PLATFORM_CARD_REQUIRED_KEYS, "platform.card")
    found_forbidden, key = has_key_recursive(card_json["card"], PLATFORM_FORBIDDEN_KEYS)
    assert not found_forbidden, f"Platform card exposed internal field: {key}"

    pdf = client.post("/v1/export/stone/pdf", json={
        "language": "BILINGUAL",
        "stone_result": analyzed_json["result"],
    })
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert "kurgin_stone_report.pdf" in pdf.headers.get("content-disposition", "")
    assert len(pdf.content) > 1000


def main():
    test_excel_workbook_schema()
    test_analysis_package_schema()
    test_api_schema_shapes()
    print("Analyzer output schema lock test OK")
    print({
        "excel_sheets": EXPECTED_SHEETS,
        "results_required_columns": len(RESULTS_REQUIRED_COLUMNS),
        "details_required_columns": len(DETAILS_REQUIRED_COLUMNS),
        "issues_required_columns": len(ISSUES_REQUIRED_COLUMNS),
        "api_endpoints_checked": [
            "/v1/health",
            "/v1/ready",
            "/v1/analyze/stone",
            "/v1/analyze/batch/json",
            "/v1/platform/stone-card",
            "/v1/export/stone/pdf",
        ],
    })


if __name__ == "__main__":
    main()
